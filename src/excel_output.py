from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models import Completeness, EngagementResult

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
PARTIAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GAP_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, num_cols: int):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws, min_width=12, max_width=60):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            if cell.value:
                lines = str(cell.value).split("\n")
                max_line = max(len(line) for line in lines)
                max_len = max(max_len, max_line)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def generate_workbook(result: EngagementResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    _write_summary_tab(wb, result)
    _write_app_tabs(wb, result)
    _write_coverage_tab(wb, result)
    _write_gaps_tab(wb, result)
    _write_observations_tab(wb, result)

    wb.save(str(output_path))
    return output_path


def _write_summary_tab(wb: Workbook, result: EngagementResult):
    ws = wb.active
    ws.title = "Summary"

    ws.append(["Engagement Summary"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Engagement", result.engagement_name])
    ws.append(["Applications Reviewed", len(result.app_results)])
    ws.append(["Total Questions", len(result.question_framework.questions)])

    total = len(result.coverage_matrix)
    full = sum(1 for c in result.coverage_matrix if c.completeness == Completeness.FULL)
    partial = sum(1 for c in result.coverage_matrix if c.completeness == Completeness.PARTIAL)
    gaps = sum(1 for c in result.coverage_matrix if c.completeness == Completeness.NOT_ADDRESSED)

    ws.append(["Fully Answered", f"{full}/{total} ({round(full / total * 100) if total else 0}%)"])
    ws.append(["Partially Answered", f"{partial}/{total}"])
    ws.append(["Not Addressed", f"{gaps}/{total}"])
    ws.append([])

    if result.summary:
        ws.append(["Executive Summary"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append([result.summary])
        ws.cell(row=ws.max_row, column=1).alignment = WRAP_ALIGNMENT

    ws.append([])
    ws.append(["Applications"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    for app in result.app_results:
        ws.append([
            app.application_name,
            f"Interviewees: {', '.join(app.interviewees) or 'N/A'}",
            f"{len(app.answered_questions)} answers",
            f"{len(app.findings)} findings",
        ])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50


def _write_app_tabs(wb: Workbook, result: EngagementResult):
    for app in result.app_results:
        tab_name = app.application_name[:31]
        ws = wb.create_sheet(title=tab_name)

        headers = [
            "Question ID",
            "Category",
            "Question",
            "Synthesized Answer",
            "Answer Type",
            "Confidence",
            "Completeness",
            "Sources",
            "Conflicts / Notes",
        ]
        ws.append(headers)
        _style_header_row(ws, len(headers))

        q_map = {q.id: q for q in result.question_framework.questions}

        for a in app.answered_questions:
            q = q_map.get(a.question_id)
            category = q.category if q else ""
            sources_str = "; ".join(
                f"{s.source_type.value}: {s.filename}" for s in a.sources
            )

            row_num = ws.max_row + 1
            ws.append([
                a.question_id,
                category,
                a.question_text,
                a.synthesized_answer,
                a.answer_type.value.title(),
                a.confidence,
                a.completeness.value.replace("_", " ").title(),
                sources_str,
                a.conflicts or "",
            ])

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGNMENT

            comp_cell = ws.cell(row=row_num, column=7)
            if a.completeness == Completeness.FULL:
                comp_cell.fill = POSITIVE_FILL
            elif a.completeness == Completeness.PARTIAL:
                comp_cell.fill = PARTIAL_FILL
            else:
                comp_cell.fill = GAP_FILL

        if app.findings:
            ws.append([])
            ws.append(["Findings"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

            finding_headers = [
                "Principle",
                "Severity",
                "Type",
                "Observation",
                "Evidence",
                "Recommendation",
            ]
            ws.append(finding_headers)
            _style_header_row(ws, len(finding_headers))

            for f in app.findings:
                ws.append([
                    f.principle_name,
                    f.severity.value.title(),
                    "Positive" if f.is_positive else "Concern",
                    f.observation,
                    f.evidence,
                    f.recommendation,
                ])

        _auto_width(ws)


def _write_coverage_tab(wb: Workbook, result: EngagementResult):
    ws = wb.create_sheet(title="Coverage Matrix")

    app_names = [r.application_name for r in result.app_results]
    headers = ["Question ID", "Category", "Question"] + app_names
    ws.append(headers)
    _style_header_row(ws, len(headers))

    coverage_map = {}
    for cell in result.coverage_matrix:
        coverage_map[(cell.question_id, cell.application_name)] = cell.completeness

    for q in result.question_framework.questions:
        row = [q.id, q.category, q.text]
        for app_name in app_names:
            comp = coverage_map.get((q.id, app_name), Completeness.NOT_ADDRESSED)
            row.append(comp.value.replace("_", " ").title())
        ws.append(row)

        row_num = ws.max_row
        for col_idx, app_name in enumerate(app_names, start=4):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            comp = coverage_map.get((q.id, app_name), Completeness.NOT_ADDRESSED)
            if comp == Completeness.FULL:
                cell.fill = POSITIVE_FILL
            elif comp == Completeness.PARTIAL:
                cell.fill = PARTIAL_FILL
            else:
                cell.fill = GAP_FILL

    _auto_width(ws)


def _write_gaps_tab(wb: Workbook, result: EngagementResult):
    ws = wb.create_sheet(title="Gaps")

    headers = ["Application", "Question ID", "Category", "Question", "Completeness", "Notes"]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    q_map = {q.id: q for q in result.question_framework.questions}

    for app in result.app_results:
        for a in app.answered_questions:
            if a.completeness in (Completeness.PARTIAL, Completeness.NOT_ADDRESSED):
                q = q_map.get(a.question_id)
                ws.append([
                    app.application_name,
                    a.question_id,
                    q.category if q else "",
                    a.question_text,
                    a.completeness.value.replace("_", " ").title(),
                    a.conflicts or "",
                ])

    _auto_width(ws)


def _write_observations_tab(wb: Workbook, result: EngagementResult):
    ws = wb.create_sheet(title="AI Observations")

    headers = [
        "Observation",
        "Evidence",
        "Source Interviews",
        "Suggested Follow-up Question",
    ]
    ws.append(headers)
    _style_header_row(ws, len(headers))

    for obs in result.ai_observations:
        ws.append([
            obs.observation,
            obs.evidence,
            ", ".join(obs.source_interviews),
            obs.suggested_followup or "",
        ])

    _auto_width(ws)
