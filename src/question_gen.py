from __future__ import annotations

import json
import time
from pathlib import Path

from src.models import AgentTrace, Question, QuestionFramework
from src.utils import DEFAULT_MODEL, count_tokens, get_anthropic_client, parse_file

SYSTEM_PROMPT = """You are an expert data audit consultant generating interview questions for a consulting engagement.

Given engagement documents (SOW, proposal, contract, project plan, transcripts, architecture docs), you will:
1. Extract the scope, systems in play, deliverables, and constraints.
2. Generate targeted interview questions organized by category.

Questions should target:
- Data usage patterns (what data, where it lives, how it flows)
- Integration patterns (how systems connect, APIs, batch processes)
- Data dependencies (upstream/downstream, timing, SLAs)
- Data quality and governance (validation, lineage, ownership)
- Pain points and known issues
- Future state and planned changes

Each question must have:
- A unique ID (format: Q-{category_abbrev}-{number}, e.g., Q-DATA-001)
- A category (e.g., "Data Sources", "Integrations", "Data Quality", "Dependencies", "Pain Points", "Future State")
- The question text
- Optional context about why the question matters

Return your response as a JSON object with this structure:
{
  "questions": [
    {
      "id": "Q-DATA-001",
      "category": "Data Sources",
      "text": "What are the primary data sources...",
      "context": "Understanding source systems helps map..."
    }
  ]
}

Generate 15-30 questions depending on engagement complexity. Prioritize questions that will reveal data architecture patterns, dependencies, and potential risks."""

QUESTION_TOOL = {
    "name": "submit_questions",
    "description": "Submit the generated interview questions",
    "input_schema": {
        "type": "object",
        "properties": {
            "questions": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "id": {"type": "string"},
                        "category": {"type": "string"},
                        "text": {"type": "string"},
                        "context": {"type": "string"},
                    },
                    "required": ["id", "category", "text"],
                },
            }
        },
        "required": ["questions"],
    },
}


def generate_questions(
    engagement_docs: list[str | Path],
    engagement_name: str = "Untitled Engagement",
) -> tuple[QuestionFramework, AgentTrace]:
    start = time.time()
    client = get_anthropic_client()

    doc_contents = []
    for doc_path in engagement_docs:
        path = Path(doc_path)
        content = parse_file(path)
        doc_contents.append(f"--- Document: {path.name} ---\n{content}")

    combined = "\n\n".join(doc_contents)
    tokens_in = count_tokens(combined)

    user_msg = (
        f"Engagement: {engagement_name}\n\n"
        f"Below are the engagement documents. Analyze them and generate targeted "
        f"interview questions.\n\n{combined}"
    )

    response = client.messages.create(
        model=DEFAULT_MODEL,
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        tools=[QUESTION_TOOL],
        tool_choice={"type": "tool", "name": "submit_questions"},
        messages=[{"role": "user", "content": user_msg}],
    )

    tool_input = None
    for block in response.content:
        if block.type == "tool_use" and block.name == "submit_questions":
            tool_input = block.input
            break

    if not tool_input:
        raise ValueError("Question generation agent did not return structured output")

    questions = [
        Question(
            id=q["id"],
            category=q["category"],
            text=q["text"],
            context=q.get("context"),
        )
        for q in tool_input["questions"]
    ]

    framework = QuestionFramework(
        engagement_name=engagement_name,
        questions=questions,
        locked=False,
        source="ai_generated",
    )

    duration = time.time() - start
    trace = AgentTrace(
        agent_name="question_generation",
        phase="A",
        input_summary=f"{len(engagement_docs)} docs, {tokens_in} tokens",
        output_summary=f"{len(questions)} questions generated",
        duration_seconds=round(duration, 2),
        llm_calls=1,
        tokens_used=response.usage.input_tokens + response.usage.output_tokens,
    )

    return framework, trace


def load_questions_from_json(filepath: str | Path) -> QuestionFramework:
    with open(filepath) as f:
        data = json.load(f)

    questions = [
        Question(
            id=q["id"],
            category=q["category"],
            text=q["text"],
            context=q.get("context"),
        )
        for q in data["questions"]
    ]

    return QuestionFramework(
        engagement_name=data.get("engagement_name", "Imported"),
        questions=questions,
        locked=data.get("locked", True),
        source="manual_upload",
    )


def load_questions_from_excel(
    filepath: str | Path,
    sheet_name: str | None = None,
    engagement_name: str = "Imported",
) -> QuestionFramework:
    import re
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    cat_pattern = re.compile(r"^\d+\.\s+(.+?)(?:\s*\(\d+\s*min\))?\s*$")
    current_category = None
    questions = []
    cat_counters: dict[str, int] = {}

    cat_abbrevs: dict[str, str] = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1, values_only=True):
        val = row[0]
        if not val or not isinstance(val, str):
            continue
        val = val.strip()
        if val.lower() in ("questions", "question"):
            continue

        match = cat_pattern.match(val)
        if match:
            current_category = match.group(1).strip()
            if current_category not in cat_abbrevs:
                words = current_category.split()
                abbrev = "".join(w[0].upper() for w in words if w[0].isalpha())[:4]
                cat_abbrevs[current_category] = abbrev
            continue

        if current_category and (val.endswith("?") or len(val) > 20):
            abbrev = cat_abbrevs.get(current_category, "GEN")
            cat_counters[abbrev] = cat_counters.get(abbrev, 0) + 1
            qid = f"Q-{abbrev}-{cat_counters[abbrev]:03d}"
            questions.append(
                Question(id=qid, category=current_category, text=val)
            )

    wb.close()

    return QuestionFramework(
        engagement_name=engagement_name,
        questions=questions,
        locked=True,
        source="manual_upload",
    )


def merge_frameworks(
    ai_framework: QuestionFramework,
    manual_questions: list[Question],
) -> QuestionFramework:
    existing_ids = {q.id for q in ai_framework.questions}
    merged = list(ai_framework.questions)

    for q in manual_questions:
        if q.id not in existing_ids:
            merged.append(q)
            existing_ids.add(q.id)

    return QuestionFramework(
        engagement_name=ai_framework.engagement_name,
        questions=merged,
        locked=False,
        source="hybrid",
    )


def lock_framework(framework: QuestionFramework) -> QuestionFramework:
    return framework.model_copy(update={"locked": True})


def export_framework_json(framework: QuestionFramework, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(framework.model_dump(), f, indent=2)
    return output_path
